"""
04_build_dataset.py — 비교공시 엑셀 원본 → data/products.json / providers.json / assumptions.json

■ 이 스크립트는 M1(문서 정규화)이 아니라 순수 데이터 변환이다. LLM이 필요 없다.
  비교공시 엑셀은 41개 사업자·324개 상품이 이미 표 형태로 정리되어 있어
  (상품설명서 PDF와 달리) 사업자마다 용어·구조가 다르지 않다.

■ 입력
  data/raw/★사전지정운용방법_상품별_비교_공시자료_2026._1분기(최종).xlsx
    시트 1 '종합'        : 상품별 원자료 (8~331행, 324개 상품)
    시트 2 '적립금_비중'  : 사업자별 위험도 구간 적립금 비중 (6~46행, 41개 사업자)

  ※ 이 파일은 openpyxl 기본 로드(load_workbook without read_only) 시
     스타일 인덱스 오류(IndexError)로 열리지 않는다. read_only=True 필수.
     파일명에 특수문자(★, 괄호)가 있어 커맨드라인 인자로 넘기면 셸에 따라
     깨지므로, data/raw/ 안을 glob으로 찾는다 — 인자를 받지 않는다.

■ 출력
  data/products.json      324개 상품 원자료 (정규화된 필드명, 결측치는 null)
  data/providers.json     41개 사업자별 위험도 비중 + 집계 적립금·상품수
  data/assumptions.json   위험도 라벨별 1년 수익률 중앙값(annual_return)과
                           같은 라벨 내 표준편차(volatility) — 실측치다.
                           주의: 이것은 '과거 실적'이지 '미래 기대수익률'이 아니다.
                           고용노동부 공시로 교체 가능한 자리이지만, 최소한
                           임의의 임시값보다는 실제 시장 분포에 근거한다.

사용법: python scripts/04_build_dataset.py
"""
import glob
import json
import statistics
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "data"

RISK_LABELS_ORDER = ["초저위험", "저위험", "중위험", "고위험"]

# 종합 시트 컬럼 순서 (1-indexed, 8행부터 데이터 시작)
PRODUCT_COLUMNS = [
    "name", "provider", "guarantee", "product_type", "risk_label", "sale_status",
    "set_date", "approval_date",
    "aum_dc", "aum_irp", "aum_total",
    "users_managed_dc", "users_managed_irp", "users_managed_total",
    "users_designated_dc", "users_designated_irp", "users_designated_total",
    "workplaces",
    "return_1m", "return_3m", "return_6m", "return_1y", "return_3y",
    "fee_pct",
]


def find_source_xlsx() -> Path:
    candidates = [
        p for p in glob.glob(str(RAW_DIR / "*비교_공시자료*.xlsx"))
        if "~$" not in p
    ]
    if not candidates:
        raise FileNotFoundError(f"{RAW_DIR}에서 비교공시 엑셀을 찾지 못했습니다.")
    return Path(candidates[0])


def clean(value):
    """'-' 등 결측 표기를 None으로. 그 외는 원값 유지."""
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in ("-", ""):
        return None
    return value


def clean_date(value) -> str | None:
    """YYYYMMDD 정수 → 'YYYY-MM-DD' 문자열."""
    value = clean(value)
    if value is None:
        return None
    s = str(int(value))
    if len(s) != 8:
        return None
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def clean_pct(value):
    """수익률·보수 값. 엑셀에 이미 %단위 숫자(2.95 = 2.95%)로 들어있으므로 그대로 둔다."""
    value = clean(value)
    if value is None:
        return None
    return round(float(value), 4)


def load_products(ws) -> list[dict]:
    products = []
    for row in ws.iter_rows(min_row=8, max_row=331, max_col=24, values_only=True):
        if row[0] is None:  # 상품명 없는 빈 행 방어
            continue
        rec = dict(zip(PRODUCT_COLUMNS, row))
        rec["guarantee"] = clean(rec["guarantee"])
        rec["product_type"] = clean(rec["product_type"])
        rec["risk_label"] = clean(rec["risk_label"])
        rec["sale_status"] = clean(rec["sale_status"])
        rec["set_date"] = clean_date(rec["set_date"])
        rec["approval_date"] = clean_date(rec["approval_date"])
        for k in ("aum_dc", "aum_irp", "aum_total",
                   "users_managed_dc", "users_managed_irp", "users_managed_total",
                   "users_designated_dc", "users_designated_irp", "users_designated_total",
                   "workplaces"):
            rec[k] = clean(rec[k])
        for k in ("return_1m", "return_3m", "return_6m", "return_1y", "return_3y", "fee_pct"):
            rec[k] = clean_pct(rec[k])
        products.append(rec)
    return products


def load_provider_risk_mix(ws) -> dict[str, dict]:
    """적립금_비중 시트: 사업자별 위험도 구간 비중(0~1 소수)."""
    mix = {}
    for row in ws.iter_rows(min_row=6, max_row=46, max_col=6, values_only=True):
        _, provider, ultra_low, low, mid, high = row
        if provider is None:
            continue
        mix[provider] = {
            "초저위험_비중": round(ultra_low, 4) if ultra_low is not None else None,
            "저위험_비중": round(low, 4) if low is not None else None,
            "중위험_비중": round(mid, 4) if mid is not None else None,
            "고위험_비중": round(high, 4) if high is not None else None,
        }
    return mix


def load_market_return_row(ws) -> dict:
    """적립금_비중 시트 3행: 라벨별 시장 전체 가중평균 1년 수익률 (교차검증용)."""
    row = next(ws.iter_rows(min_row=3, max_row=3, max_col=6, values_only=True))
    _, _, ultra_low, low, mid, high = row
    return {"초저위험": ultra_low, "저위험": low, "중위험": mid, "고위험": high}


def build_providers(products: list[dict], risk_mix: dict[str, dict]) -> list[dict]:
    by_provider: dict[str, list[dict]] = {}
    for p in products:
        by_provider.setdefault(p["provider"], []).append(p)

    providers = []
    for name, mix in risk_mix.items():
        prods = by_provider.get(name, [])
        total_aum = sum(p["aum_total"] for p in prods if p["aum_total"] is not None)
        providers.append({
            "name": name,
            "product_count": len(prods),
            "total_aum": total_aum,
            **mix,
        })
    providers.sort(key=lambda p: -(p["total_aum"] or 0))
    return providers


def label_stats(products: list[dict], field: str, exclude_zero_fee_unguaranteed: bool = False) -> dict:
    """
    exclude_zero_fee_unguaranteed: 보수(fee_pct) 통계 전용 옵션.
    원금 비보장(집합투자증권 등) 상품인데 합성총보수가 정확히 0.000%로
    공시된 경우가 소수 존재한다 — 실제 무보수가 아니라 공시 누락으로 보는
    편이 합리적이다(원금보장형만 진짜 0%가 가능). 이 옵션을 켜면 그런
    행을 통계에서 제외한다. 수익률(field='return_*') 계산에는 쓰지 않는다.
    """
    by_label: dict[str, list[float]] = {lbl: [] for lbl in RISK_LABELS_ORDER}
    for p in products:
        v = p.get(field)
        lbl = p.get("risk_label")
        if v is None or lbl not in by_label:
            continue
        if exclude_zero_fee_unguaranteed and v == 0 and p.get("guarantee") != "보장":
            continue
        by_label[lbl].append(v)

    stats = {}
    for lbl, values in by_label.items():
        if not values:
            continue
        stats[lbl] = {
            "count": len(values),
            "min": round(min(values), 2),
            "median": round(statistics.median(values), 2),
            "max": round(max(values), 2),
            "stdev": round(statistics.stdev(values), 2) if len(values) > 1 else 0.0,
        }
    return stats


def build_assumptions(return_stats: dict) -> dict:
    """
    위험도 라벨별 GRADE_ASSUMPTIONS.
    annual_return = 라벨 내 1년 수익률 중앙값 (실측, %→소수)
    volatility    = 라벨 내 1년 수익률 표준편차 (실측 분산, %→소수)

    이것은 '기대수익률 예측'이 아니라 '2026년 1분기 시점 실제로 관측된 분포'다.
    프로젝트의 핵심 메시지(같은 라벨 안에서도 성과가 크게 갈린다)를 시뮬레이션
    밴드에도 그대로 반영한다는 점에서, 임의의 임시값보다 훨씬 정직한 선택이다.
    """
    source = "사전지정운용방법 상품별 비교공시 2026년 1분기 — 실측 1년 수익률 분포(중앙값·표준편차). 미래 기대수익률이 아니라 과거 실적 기반 근사치."
    assumptions = {}
    for lbl in RISK_LABELS_ORDER:
        s = return_stats.get(lbl)
        if not s:
            continue
        assumptions[lbl] = {
            "annual_return": round(s["median"] / 100, 4),
            "volatility": round(s["stdev"] / 100, 4),
            "source": source,
            "as_of": "2026-03-31",
        }
    return assumptions


def print_report(products: list[dict], providers: list[dict],
                  return_stats: dict, fee_stats: dict, market_returns: dict):
    total_aum = sum(p["aum_total"] for p in products if p["aum_total"] is not None)
    ultra_low_aum = sum(p["aum_total"] for p in products
                         if p["aum_total"] is not None and p["risk_label"] == "초저위험")

    print("=" * 76)
    print(f"상품 {len(products)}개 · 사업자 {len(providers)}개")
    print(f"총 적립금 {total_aum / 1e12:.1f}조 원 · 초저위험 비중 "
          f"{ultra_low_aum / total_aum * 100:.1f}% ({ultra_low_aum / 1e12:.1f}조 원)")

    print("\n[발견①] 라벨 내 1년 수익률 격차")
    print(f"{'라벨':6s} {'상품수':>5s} {'최저':>8s} {'중앙값':>8s} {'최고':>8s} {'격차':>8s}")
    for lbl in RISK_LABELS_ORDER:
        s = return_stats.get(lbl)
        if not s:
            continue
        spread = s["max"] - s["min"]
        print(f"{lbl:6s} {s['count']:5d} {s['min']:7.2f}% {s['median']:7.2f}% "
              f"{s['max']:7.2f}% {spread:7.2f}%p")

    print("\n[발견②] 라벨 구간 겹침")
    present = [lbl for lbl in RISK_LABELS_ORDER if lbl in return_stats]
    for lower, higher in zip(present, present[1:]):
        lo, hi = return_stats[lower], return_stats[higher]
        overlap = lo["max"] >= hi["min"]
        mark = "겹침" if overlap else "겹침 없음"
        print(f"  {lower} 최고 {lo['max']:.2f}% vs {higher} 최저 {hi['min']:.2f}% → {mark}")

    print("\n[발견③] 라벨 내 보수 배수")
    print(f"{'라벨':6s} {'최저':>8s} {'중앙값':>8s} {'최고':>8s} {'배수':>8s}")
    for lbl in RISK_LABELS_ORDER:
        s = fee_stats.get(lbl)
        if not s or s["min"] == 0:
            continue
        mult = s["max"] / s["min"]
        print(f"{lbl:6s} {s['min']:7.3f}% {s['median']:7.3f}% {s['max']:7.3f}% {mult:7.1f}배")

    print("\n[교차검증] 적립금_비중 시트의 라벨별 시장 평균 1년 수익률 vs 종합 시트 중앙값")
    for lbl in RISK_LABELS_ORDER:
        mkt = market_returns.get(lbl)
        med = return_stats.get(lbl, {}).get("median")
        if mkt is None or med is None:
            continue
        print(f"  {lbl:6s} 시장평균(비중가중) {mkt*100:6.2f}%  vs  중앙값(단순) {med:6.2f}%")


def main():
    src = find_source_xlsx()
    print(f"입력: {src}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    products = load_products(wb["종합"])
    risk_mix = load_provider_risk_mix(wb["적립금_비중"])
    market_returns = load_market_return_row(wb["적립금_비중"])
    providers = build_providers(products, risk_mix)

    return_stats = label_stats(products, "return_1y")
    fee_stats = label_stats(products, "fee_pct", exclude_zero_fee_unguaranteed=True)
    assumptions = build_assumptions(return_stats)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "products.json").write_text(
        json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "providers.json").write_text(
        json.dumps(providers, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "assumptions.json").write_text(
        json.dumps(assumptions, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n생성: data/products.json ({len(products)}건)")
    print(f"생성: data/providers.json ({len(providers)}건)")
    print(f"생성: data/assumptions.json ({len(assumptions)}개 라벨)")
    print()
    print_report(products, providers, return_stats, fee_stats, market_returns)


if __name__ == "__main__":
    main()
